import numpy as np
import pandas as pd
import openpyxl
import time
import csv
import sys
import os

path = "C:\\Users\\user\\Desktop\\台股大盤資訊"

os.chdir(path)                                  # 輸入指定路徑，修改當前目錄
os.getcwd()                                     # 獲取目錄路徑

def Least_square_method(data):
    if len(data) > 6:
        data = data[-7:-1]
    x = np.arange(0,len(data))
    y = np.array(data)
    N = len(y)
    B = (sum(x[i] * y[i] for i in x) - 1./N*sum(x)*sum(y)) / (sum(x[i]**2 for i in x) - 1./N*sum(x)**2)
    A = 1.*sum(y)/N - B * 1.*sum(x)/N
    return int(A + B * N)

def Start_yy_mm(year,month):
    matrix = []
    x = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    yy    = int(x[:4]) - 1911
    mm    = int(x[5:7])
    year  = int(year)
    month = int(month)
    while year != yy or month != mm:
        if month < 10:
            matrix.append([str(year),'0' + str(month)])
        else:
            matrix.append([str(year),str(month)])
        month += 1
        if month > 12:
            year += 1
            month = 1
    if month < 10:
        matrix.append([str(year),'0' + str(month)])
    else:
        matrix.append([str(year),str(month)])
    return matrix
    


name = input('請輸入你的檔案名稱:')
try:
    wb = openpyxl.load_workbook(name + '.xlsx')
    establish = False
except Exception as X:
    print('找不到該檔案')
    response = input('請問你要新建一個檔案嗎?\n回答:')
    if response == 'yes':
        wb = openpyxl.Workbook()
        wb.save(name + '.xlsx')
        openpyxl.load_workbook(name + '.xlsx')
        establish = True
    else:
        sys.exit()

if establish:
    year  = input('請輸入你要的年份:')
    month = input('請輸入你要的月份:')
    total = Start_yy_mm(year,month)

    share_price = []
    discriminate = ['賣出']
    index = [0,0,0]

    sheet = wb['Sheet']
    sheet['A1'] = '日期'
    sheet['B1'] = '收盤價'
    sheet['C1'] = '預測價'
    sheet['D1'] = '判定'
    sheet['E1'] = '訊號'

    for t in total:
        date = []
        data = dict()
        if os.path.isfile(path + '\\' + t[0] + t[1] + '.csv'):
            pass
        else:
            print(t[0] + t[1] + '.csv' + "這個檔案不存在，請重新輸入。")
            sys.exit()
        with open(t[0] + t[1] + '.csv', newline='') as csvfile:
            rows = csv.reader(csvfile)
            loop_number = 0
            for row in rows:
                loop_number += 1
                if loop_number < 3:
                    continue
                date.append(row[0])
                price = float(row[4].replace(',',''))
                day   = row[0]
                data[day] = price
        x = sheet.max_row + 1
        for d in date:
            share_price.append(float(data[d]))
            sheet['A'+str(x)] = d
            sheet['B'+str(x)] = int(data[d])
            if len(share_price) > 6:
                sheet['C'+str(x)] = int(Least_square_method(share_price))
                sheet['D'+str(x)] = int(data[d]) - int(Least_square_method(share_price))
                index.append(int(data[d]) - int(Least_square_method(share_price)))
                if (index[-2] < 0 < index[-1] or 0 < index[-2] < index[-1]) or (index[-3] < 0 and index[-2] < index[-1] and -30 < index[-1] < 0):
                    if discriminate[-1] == '買進':
                        pass
                    else:
                        sheet['E'+str(x)] = '買進'
                        discriminate.append('買進')
                elif (index[-2] > 0 > index[-1] or 0 > index[-2] > index[-1]) or (index[-3] > 0 and index[-2] > index[-1] and 0 < index[-1] < 30):
                    if discriminate[-1] == '賣出':
                        pass
                    else:
                        sheet['E'+str(x)] = '賣出'
                        discriminate.append('賣出')
            x += 1
        wb.save(name + '.xlsx')
else:
    df = pd.read_excel(name + '.xlsx')
    share_price = [df.loc[len(df)-6]['收盤價'],
                   df.loc[len(df)-5]['收盤價'],
                   df.loc[len(df)-4]['收盤價'],
                   df.loc[len(df)-3]['收盤價'],
                   df.loc[len(df)-2]['收盤價'],
                   df.loc[len(df)-1]['收盤價']]
    index = [df.loc[len(df)-3]['判定'],
             df.loc[len(df)-2]['判定'],
             df.loc[len(df)-1]['判定']]
    df2 = df.loc[:,['訊號']].dropna()
    discriminate = [str(df2.tail(1))[11:13]]

    total = Start_yy_mm(df.loc[len(df)-1]['日期'][0:3],df.loc[len(df)-1]['日期'][4:6])
    sheet = wb['Sheet']

    for t in total:
        date = []
        data = dict()
        if os.path.isfile(path + '\\' + t[0] + t[1] + '.csv'):
            pass
        else:
            print(t[0] + t[1] + '.csv' + "這個檔案不存在，請重新輸入。")
            sys.exit()
        with open(t[0] + t[1] + '.csv', newline='') as csvfile:
            rows = csv.reader(csvfile)
            loop_number = 0
            for row in rows:
                loop_number += 1
                if loop_number < 3 or df.loc[len(df)-1]['日期'] >= row[0]:
                    continue
                date.append(row[0])
                price = float(row[4].replace(',',''))
                day   = row[0]
                data[day] = price
        x = sheet.max_row + 1
        for d in date:
            share_price.append(float(data[d]))
            sheet['A'+str(x)] = d
            sheet['B'+str(x)] = int(data[d])
            if len(share_price) > 6:
                sheet['C'+str(x)] = int(Least_square_method(share_price))
                sheet['D'+str(x)] = int(data[d]) - int(Least_square_method(share_price))
                index.append(int(data[d]) - int(Least_square_method(share_price)))
                if (index[-2] < 0 < index[-1] or 0 < index[-2] < index[-1]) or (index[-3] < 0 and index[-2] < index[-1] and -30 < index[-1] < 0):
                    if discriminate[-1] == '買進':
                        pass
                    else:
                        sheet['E'+str(x)] = '買進'
                        discriminate.append('買進')
                elif (index[-2] > 0 > index[-1] or 0 > index[-2] > index[-1]) or (index[-3] > 0 and index[-2] > index[-1] and 0 < index[-1] < 30):
                    if discriminate[-1] == '賣出':
                        pass
                    else:
                        sheet['E'+str(x)] = '賣出'
                        discriminate.append('賣出')
            x += 1
        wb.save(name + '.xlsx')


df = pd.read_excel(name + '.xlsx')
response = input('最新的日期為' + df.loc[len(df)-1]['日期'] + '，請問是否需要手動輸入:')
if response == 'yes':
    share_price = [df.loc[len(df)-6]['收盤價'],
                   df.loc[len(df)-5]['收盤價'],
                   df.loc[len(df)-4]['收盤價'],
                   df.loc[len(df)-3]['收盤價'],
                   df.loc[len(df)-2]['收盤價'],
                   df.loc[len(df)-1]['收盤價']]
    index = [df.loc[len(df)-3]['判定'],
             df.loc[len(df)-2]['判定'],
             df.loc[len(df)-1]['判定']]
    df2 = df.loc[:,['訊號']].dropna()
    discriminate = [str(df2.tail(1))[11:13]]

    date  = input('請輸入最新日期(格式為 xxx/xx/xx ):')
    price = int(input('請輸入最新日期的收盤價:'))

    wb = openpyxl.load_workbook(name + '.xlsx')
    sheet = wb['Sheet']
    x = sheet.max_row + 1
    sheet['A'+str(x)] = date
    sheet['B'+str(x)] = price
    sheet['C'+str(x)] = int(Least_square_method(share_price))
    sheet['D'+str(x)] = price - int(Least_square_method(share_price))
    index.append(price - int(Least_square_method(share_price)))
    if (index[-2] < 0 < index[-1] or 0 < index[-2] < index[-1]) or (index[-3] < 0 and index[-2] < index[-1] and -30 < index[-1] < 0):
        if discriminate[-1] == '買進':
            pass
        else:
            sheet['E'+str(x)] = '買進'
            discriminate.append('買進')
    elif (index[-2] > 0 > index[-1] or 0 > index[-2] > index[-1]) or (index[-3] > 0 and index[-2] > index[-1] and 0 < index[-1] < 30):
        if discriminate[-1] == '賣出':
            pass
        else:
            sheet['E'+str(x)] = '賣出'
            discriminate.append('賣出')
    wb.save(name + '.xlsx')
else:
    sys.exit()
